from __future__ import annotations
from itertools import zip_longest
from logging import getLogger
from typing import Any, Literal, TYPE_CHECKING
import numpy as np
try:
    from openpyxl.worksheet.worksheet import Worksheet
    from openpyxl.chart import ScatterChart, Reference, Series
except ImportError as exc:
    raise ImportError("openpyxl is not installed. ") from exc
from exptoolkit.plotter.backends._base import Target
from exptoolkit.plotter.colors import parse_color

if TYPE_CHECKING:
    # import Series object directly since openpyxl.chart.Series does not work
    from openpyxl.chart.series import Series as _Series

logger = getLogger(__name__)


class OpenPyXlTarget(Target):
    """OpenPyXlTarget backend for plotting graphs. Implements the Target protocol."""
    def __init__(
        self,
        ws: Worksheet,
        chart: ScatterChart | None = None,
        addr_default: str = "E5",
        ):
        """Initialize the OpenPyXlTarget with an openpyxl Worksheet object.
        Args:
            ws (Worksheet): An openpyxl Worksheet object where the plots will be drawn.
            chart (ScatterChart | None): An openpyxl chart object.
                If None, a ScatterChart will be created by default.
            addr_default (str): The default cell address where the chart will be placed
                if a new chart is created.
        """
        self.ws = ws
        self.chart = chart or ScatterChart()
        if chart is None:
            self.ws.add_chart(self.chart, addr_default)

    def add_line(self, x, y, color=None, label=None, **kwargs):
        max_col = self.ws.max_column
        prefix = label + "_" if label else ""
        self.ws.cell(row=1, column=max_col + 1).value = prefix + getattr(x, "name", "x")
        self.ws.cell(row=1, column=max_col + 2).value = prefix + getattr(y, "name", "y")
        x_np = np.asarray(x)
        y_np = np.asarray(y)
        for i, (x_val, y_val) in enumerate(zip_longest(x_np, y_np, fillvalue=np.nan)):
            self.ws.cell(row=i + 2, column=max_col + 1).value = x_val
            self.ws.cell(row=i + 2, column=max_col + 2).value = y_val
        x_ref = Reference(self.ws, min_col=max_col + 1, min_row=2, max_row=1 + len(x_np))
        y_ref = Reference(self.ws, min_col=max_col + 2, min_row=2, max_row=1 + len(y_np))
        series: _Series = Series(y_ref, x_ref, title=label)
        if color is not None:
            c = parse_color(color).as_hex()[1:]
            series.graphicalProperties.line.solidFill = c
        self.chart.series.append(series)
        return series

    def add_scatter(self, x, y, c=None, color=None, label=None, color_scale="linear", **kwargs):
        if c is not None:
            logger.warning("Color mapping for scatter plots is not supported in Openpyxl backend."
                " Ignoring color information.")
        max_col = self.ws.max_column
        prefix = label + "_" if label else ""
        self.ws.cell(row=1, column=max_col + 1).value = prefix + getattr(x, "name", "x")
        self.ws.cell(row=1, column=max_col + 2).value = prefix + getattr(y, "name", "y")
        x = np.asarray(x)
        y = np.asarray(y)
        for i, (x_val, y_val) in enumerate(zip_longest(x, y, fillvalue=np.nan)):
            self.ws.cell(row=i + 2, column=max_col + 1).value = x_val
            self.ws.cell(row=i + 2, column=max_col + 2).value = y_val
        x_ref = Reference(self.ws, min_col=max_col + 1, min_row=2, max_row=1 + len(x))
        y_ref = Reference(self.ws, min_col=max_col + 2, min_row=2, max_row=1 + len(y))
        series: _Series = Series(y_ref, xvalues=x_ref, title=label)
        series.graphicalProperties.line.noFill = True
        series.marker.symbol = 'circle'
        if color is not None:
            color_obj = parse_color(color)
            series.marker.spPr.solidFill = color_obj.as_hex()[1:]
            series.marker.spPr.ln.solidFill = color_obj.as_hex()[1:]
        self.chart.series.append(series)

        return series

    def set_ax_label(self, axis: Literal["x", "y"], label: str) -> None:
        if axis == "x":
            self.chart.x_axis.title = label
        elif axis == "y":
            self.chart.y_axis.title = label
        else:
            raise ValueError(f"Invalid axis: {axis}. Axis must be 'x' or 'y'.")

    def set_scale(self, axis: Literal["x", "y"], scale: Literal["linear", "log"]) -> None:
        if axis == "x":
            self.chart.x_axis.scaling.logBase = 10 if scale == "log" else None
        elif axis =="y":
            self.chart.y_axis.scaling.logBase = 10 if scale == "log" else None
        else:
            raise ValueError("Invalid axis: {axis}. Axis must bet 'x' or 'y'.")

    def set_title(self, title: str) -> None:
        self.chart.title = title

    def set_aspect(self, aspect: Literal["equal", "auto"]) -> None:
        logger.warning("Setting aspect ratio is not supported in Openpyxl backend. Skipping.")

    def reverse_axis(self, x: bool | None = None, y: bool | None = None) -> None:
        if x is not None:
            self.chart.x_axis.scaling.orientation = 'maxMin' if x else 'minMax'
        if y is not None:
            self.chart.y_axis.scaling.orientation = 'maxMin' if y else 'minMax'

    @classmethod
    def from_obj(cls, obj: Any):
        """Create an OpenpyxlTarget from an openpyxl Worksheet or Chart object."""
        if isinstance(obj, OpenPyXlTarget):
            return obj
        if isinstance(obj, Worksheet):
            return cls(ws=obj)
        if isinstance(obj, tuple) and len(obj) == 2:
            ws, chart = obj
            if not isinstance(ws, Worksheet):
                raise TypeError("First element of the tuple must be an openpyxl Worksheet.")
            if not isinstance(chart, ScatterChart):
                raise TypeError("Second element of the tuple must be "
                                "an openpyxl ScatterChart.")
            return cls(ws=ws, chart=chart)
        raise TypeError(
            "Unrecognized object type for OpenPyXlTarget. "
            f"Expected (Worksheet, ScatterChart) tuple or Worksheet. "
            f"Got: {type(obj).__name__}"
            )
